import openpyxl 
from email.message import EmailMessage
import ssl
import smtplib

FILE_PATH = 'registru.xlsx'

SENDER = 'alexurluescu23@gmail.com'


def open_xlsx(path:str) -> openpyxl.load_workbook:
    """ Function that opens a excel file from a deticated path"""
    book = openpyxl.load_workbook(path, data_only=True)
    return book


def get_table_data(xls_file:openpyxl.load_workbook) -> list:
    """ Function that returns data from a sheet section """
    data_array = []
    active_sheet = xls_file.active
    data_cells = active_sheet['A4':'D5']
    for row in data_cells:
        data_array.append([cell_data.value for cell_data in row])

    return data_array


def send_emails(data: list, email_sender):

    for student in data:
        student_firstname = student[0]
        student_lastname = student[1]
        student_email = student[3]
        student_grade = str(student[2])

        email_subject = 'Nota examen Fizica'
        email_body = 'Buna ziua ' + student_firstname + " " + student_lastname + ', ai luat nota ' + student_grade + "."


        email_emisor = email_sender
        email_parola = 'xecgxqiixecxaayw'
        email_receptor = student_email


        email = EmailMessage()
        email['From'] = email_emisor
        email['To'] = email_receptor
        email['Subject'] = email_subject
        email.set_content(email_body)

        contexto = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=contexto) as smtp:
            smtp.login(email_emisor, email_parola)
            smtp.sendmail(email_emisor, email_receptor, email.as_string())


def main():
    
    book = open_xlsx(FILE_PATH)
    data = get_table_data(book)
    send_emails(data, SENDER)
    

if __name__ == "__main__":
    main()

