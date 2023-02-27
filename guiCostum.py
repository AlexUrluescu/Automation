
from email.message import EmailMessage
import ssl
import smtplib
from customtkinter import CTk, CTkButton, CTkCheckBox, CTkComboBox,CTkEntry, CTkLabel, CTkFrame, CTkTextbox, CTkProgressBar
from tkinter import *
from tkinter import filedialog
import openpyxl
import time
from tkinter import ttk
import logging
import customtkinter
import os
import os.path
import webbrowser
import xlsxwriter

# customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

# customtkinter.set_appearance_mode("dark")
# customtkinter.set_appearance_mode("light")

logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")


root = CTk()
root.geometry("620x750")
root.title("Send students grades v1.0.0")
root.resizable(0,0)

drop_down_sheet_list = StringVar()
drop_down_sheet_list.set("SheetName")
contor_students = 0

grades_var = IntVar()
absente_var = IntVar()
extra_text_var = IntVar()
restante_var = IntVar()
remember_var = IntVar()

numar = 0
value = 0

costumize_var = True
info_var = True

PATH = "./email_file.txt"
PATH_PASSWORD = "./password_file.txt"

# -------------- fecth email -----------------------------------------------

date_frame = CTkFrame(root, height=60)
date_frame.grid(row = 2, column=0, sticky="we")

input_email = CTkEntry(date_frame, width=250, border_color="#C0C0C0")
input_email.grid(column=1, row=0, padx=10, pady=30, sticky="w")

input_password = CTkEntry(date_frame, border_color="white")
input_password.grid(column=3, row=0, padx=10, pady=10, sticky="w")

# ------------- creare fisier_email.txt in caz ca nu exista --------------

if(os.path.isfile(PATH) == False):
    file = open("email_file.txt", "x")

    logging.debug(f"Create {PATH}")
    
# ---- citire date din fisier_email.txt ---------
file = open("email_file.txt", "r")
email = file.read()
input_email.delete(0,END)
input_email.insert(0,email)
file.close()

logging.debug(f"Read {PATH}")

# ------------- creare password.txt in caz ca nu exista --------------

if(os.path.isfile(PATH_PASSWORD) == False):
    file = open("password_file.txt", "x")

    logging.debug(f"Create {PATH_PASSWORD}")

# ---- citire date din password.txt ---------
file = open("password_file.txt", "r")
password = file.read()
input_password.delete(0,END)
input_password.insert(0,password)

file.close()

logging.debug(f"Read {PATH_PASSWORD}")
# ----------------------------------------------------------------------

if(remember_var.get() == 0):
    logging.debug("Is inactiv")

# ----- functii ---------------
def save_email(email:str):
    if(remember_var.get() == 1):
        file = open("email_file.txt", "w")
        file.write(email)

        logging.debug(f"Writed in {PATH}")

    if(remember_var.get() == 0):
        file = open("email_file.txt", "w")
        file.write("")
    

        logging.debug("The email was not saved")


def save_password(password:str):
    if(remember_var.get() == 1):
        file = open("password_file.txt", "w")
        file.write(password)

        logging.debug(f"Writed in {PATH_PASSWORD}")

    if(remember_var.get() == 0):
        # os.remove("password.txt")
        file = open("password_file.txt", "w")
        file.write("")
    
        logging.debug("The password was not saved")


def template_btn_clb():

    workbook = xlsxwriter.Workbook('Template.xlsx', {'constant_memory': False})
    worksheet = workbook.add_worksheet()

    data =[
        ["Popescu", "Ionut", "popescuionut@gmail.com", 9, 2, "nu"],
    ]

    options = {
        "data": data,
        "columns":[{'header': 'FirstName'},
                   {'header': 'LastName'},
                   {'header': 'Email'},
                   {'header': 'Grades'},
                   {'header': 'Absente'},
                   {'header': 'Restante'},]
    }

    worksheet.add_table('B3:G4', options)

    workbook.close()

    logging.debug("The template excel file was created")


def costumize():
    global costumize_var

    if(costumize_var == True):
        logging.debug("It's True")

        customtkinter.set_appearance_mode("light")
        
        mode_button.configure(text="Dark", fg_color="black", hover_color="black", text_color="white")
        path_to_file.configure(text_color="black")
        sheet_label.configure(text_color="black")
        email_label.configure(text_color="black")
        subject_label.configure(text_color="black")
        grades_checkbutton.configure(text_color="black")
        absente_checkbutton.configure(text_color="black")
        extraText_checkbutton.configure(text_color="black")
        restante_checkbutton.configure(text_color="black")
        date_label.configure(text_color="black")
        room_label.configure(text_color="black")
        details_author_label.configure(text_color="black")
        progress_bar_label.configure(text_color="black")
        label_info.configure(text_color="black")
        rememberMe_checkbutton.configure(text_color="black")

        costumize_var = False
    
    else:
        logging.debug("It's False")

        customtkinter.set_appearance_mode("dark")

        mode_button.configure(text="Light", fg_color="white", hover_color="white", text_color="black")
        path_to_file.configure(text_color="white")
        sheet_label.configure(text_color="white")
        email_label.configure(text_color="white")
        subject_label.configure(text_color="white")
        grades_checkbutton.configure(text_color="white")
        absente_checkbutton.configure(text_color="white")
        extraText_checkbutton.configure(text_color="white")
        restante_checkbutton.configure(text_color="white")
        date_label.configure(text_color="white")
        room_label.configure(text_color="white")
        details_author_label.configure(text_color="white")
        progress_bar_label.configure(text_color="white")
        label_info.configure(text_color="white")
        rememberMe_checkbutton.configure(text_color="white")

        costumize_var = True


def file_select_btn_clb():
    root.filename = filedialog.askopenfilename(initialdir="/", title="Select a file", filetypes = (("Excel files", "*.xlsx"),("All files", "*.*")))
    path_label.configure(text=root.filename)
    path = root.filename
    sheets_btn.configure(state=NORMAL)

    logging.debug(path)
    return path


def parameters_selected_btn_clb():
    path = (path_label.cget("text"))

    logging.debug(path)
    return path


def open_xlsx(path:str) -> openpyxl.load_workbook:
    book = openpyxl.load_workbook(path, data_only=True)
    root.update()

    logging.debug(book)
    return book


def get_sheet(book:str) -> list:
    list_data = []
    sheet = drop_down_sheet_list.get()
    book.active = book[sheet]
    book_active = book.active
    logging.debug(book_active.tables.items())
    paramas_table = book_active.tables.items()
    for data in paramas_table:
        list_data.append(data[1])

    list_data.append(book_active)
    
    return list_data


def email_emisor() -> str:
    email = input_email.get().strip()
    return email


def password_emisor() -> str:
    password = input_password.get()
    return password


def get_subject() -> str:
    subject = input_subject.get().capitalize()
    return subject


def load_sheets_clb():
    combobox.configure(state=NORMAL)
    path = parameters_selected_btn_clb()
    logging.debug(path)
    book = open_xlsx(path)
    logging.debug(book)
    combobox.configure(values=book.sheetnames)
    logging.debug(book.sheetnames)


def extraText_callback():
    if(extra_text_var.get() == 1):
        text_proba.grid(row=1, column=0, padx=10, pady=10, sticky="w")

    if(extra_text_var.get() == 0):
        text_proba.place(x=1000, y=1000)


def restante_callback():
    if(restante_var.get() == 1):
        restante_frame.grid(row=4, column=0, sticky="we")
        input_date.configure(state=NORMAL)
        input_room.configure(state=NORMAL)

        absente_checkbutton.configure(state=DISABLED)
        grades_checkbutton.configure(state=DISABLED)
        grades_checkbutton.deselect()
        absente_checkbutton.deselect()
        

    if(restante_var.get() == 0):
        restante_frame.place(x=1000, y=1000)
        
        absente_checkbutton.configure(state=NORMAL)
        grades_checkbutton.configure(state=NORMAL)
        grades_checkbutton.select()


def clear_widget(contor_students):
    info_students_data.configure(state=NORMAL)
    info_students = f"Found: {contor_students} students"
    info_students_data.delete(1.0, END)
    info_students_data.insert(1.0, info_students)
    info_students_data.configure(state=DISABLED)
    

def get_table_data(sheet:object, range:str) -> list:
    global contor_students
    data_array = []
    array_students = []

    active_sheet = sheet
    logging.debug(active_sheet)
    logging.debug(f"Aici este {range}")

    data_cells = active_sheet[range]
    for row in data_cells:
        data_array.append([cell_data.value for cell_data in row])
   

    for student in data_array:
        if(student[0] == None):
            break

        array_students.append(student)

    array_students.pop(0)

    logging.debug(array_students)

    
    contor_students = len(array_students)
    clear_widget(contor_students)
    logging.debug(contor_students)
    parameters_selected.configure(state=NORMAL)
    return array_students


def get_data_btn_clb():
    path = parameters_selected_btn_clb()
    book = open_xlsx(path)
    range,sheet = get_sheet(book)
    logging.debug(f"Range {range}, Sheet {sheet}")
    email = email_emisor()
    logging.debug(email)
    password = password_emisor()
    logging.debug(password)
    get_table_data(sheet, range)
    

def get_room() -> str:
    room = input_room.get()
    return room


def get_date() -> str:
    date = input_date.get()
    return date


def send_emails(data:list, email:str, password:str, subject:str, room:str, date:str):
    global numar
    global value

    logging.debug(f"from send_emails_data -> {data}")

    if(email == "" or password == "" or subject == ""):
        error_label.grid(row=0, column=2, padx=5)
        logging.debug("The entries are not complete")

        return
    

    error_label.place(x=2000, y=2000)

    # progress bar placement in the root ----------------
    progress_bar_label.pack(padx=10)
    progress_bar.pack(padx=10, pady=10) 
    

    range_progress = len(data) # variable for the dynamic increment
   
    # ------------------------------------------------------

   

    for student in data:

        student_firstName = student[0]
        student_lastName = student[1]
        student_email = student[2]
        student_grade = student[3]
        student_abs = student[4]
        student_restanta = student[5]

        email_emisor = email
        email_password = password
        email_receptor = student_email

        if(grades_var.get() == 1 and absente_var.get() == 0):
            email_subject = 'Nota examen la ' + subject
            email_body = 'Buna ziua ' + student_firstName + " " + student_lastName + ', ai luat nota ' + str(student_grade) + "\n" + text_proba.get(1.0, END) + "."

        elif(absente_var.get() == 1 and grades_var.get() == 0):
            email_subject = 'Numarul de absente la ' + subject
            email_body = 'Buna ziua ' + student_firstName + " " + student_lastName + ', ai in total ' + str(student_abs) + " absente. \n" + text_proba.get(1.0, END) + "."

        elif(absente_var.get() == 1 and grades_var.get() == 1):
            email_subject = 'Notele si absentele la ' + subject
            email_body = 'Buna ziua ' + student_firstName + " " + student_lastName + ', ai nota ' + str(student_grade) + ' si ai in total ' + str(student_abs) + " absente. \n" + text_proba.get(1.0, END) + "."

        if(restante_var.get() == 1):
            if(int(student_grade) < 5):
                email_subject = "Informatii despre restanta la " + subject
                email_body = "Buna ziua " + student_firstName + " " + student_lastName + ", ai restanta in data de " + date +" in clasa " + room + ".\n" + text_proba.get(1.0, END) + "."

            if(int(student_grade) > 5):
                email_subject = "Informatii despre restanta la " + subject
                email_body = "Buna ziua " + student_firstName + " " + student_lastName + ", esti integralist, felicitari!.\n" + text_proba.get(1.0, END) + "."

        em = EmailMessage()
        em['From'] = email_emisor
        em['To'] = email_receptor
        em['Subject'] = email_subject
        em.set_content(email_body)

        contexto = ssl.create_default_context()

        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=contexto) as smtp:
            smtp.login(email_emisor, email_password)
            smtp.sendmail(email_emisor, email_receptor, em.as_string())

        #progress bar dynamic increment ----------------------------------
       
       
        progress_bar['value'] += 100/range_progress
        progress_bar_label.configure(text= str(round(progress_bar['value'], 1))+"%")
        root.update_idletasks()
        time.sleep(1)

        # --------------------------------------------------------------------

        logging.debug(f'Sended to {student_firstName} {student_lastName}')



def tutorial_button_info_clb():
    webbrowser.open_new_tab("https://www.youtube.com/watch?v=DDVpKvJXRz8&list=PLdkIA_6OrXkLsQFuORCmRnyAEhO4Niiux&index=1")

    logging.debug("link apasat")


def button_info_clb():
    global info_var

    if(info_var):
        frame_info.grid(row=0, column=1, pady=10)
        button_info.configure(text="Hide")

        info_var = False
    
    else:
        frame_info.place(x=2000, y=2000)
        button_info.configure(text="Info password")

        info_var = True   
    
    
def main():
    path = parameters_selected_btn_clb()
    book = open_xlsx(path)
    range,sheet = get_sheet(book)
    email = email_emisor()
    password = password_emisor()
    subject = get_subject()
    room = get_room()
    date = get_date()
    data = get_table_data(sheet, range)
    send_emails(data, email, password, subject, room, date)
    save_email(email)
    save_password(password)

# ---------- Frames --------------------------------------------------

path_frame = CTkFrame(root, height=70)
path_frame.grid(row = 0, column = 0, sticky="we")

sheet_frame = CTkFrame(root, height=60)
sheet_frame.grid(row = 1, column=0, sticky="we")

checkbuttons_frame = CTkFrame(root, height=60)
checkbuttons_frame.grid(row = 3, column=0, sticky="we")

restante_frame = CTkFrame(root)

buttons_frame = CTkFrame(root)
buttons_frame.grid(row=5, column=0, sticky="we")

textBox_Frame = CTkFrame(root)
textBox_Frame.grid(row=6, column=0, sticky="we")

progressBar_frame = CTkFrame(root, height=75)
progressBar_frame.grid(row=7, column=0, sticky="we")

block_frame = CTkFrame(root)
block_frame.grid(row=8, column=0, sticky="we")

contact_details_frame = CTkFrame(root)
contact_details_frame.place(x=100, y=715)
# contact_details_frame.pack()

# ------------------------------------------------------------------------

# ------------- Content Path Frame ------------------------------------------------

path_to_file = CTkLabel(path_frame, text="Path", font=('Helvetica', 18), text_color="white")
path_to_file.grid(column = 0, row = 0, padx=10, pady=10)

path_label = CTkLabel(path_frame, text="", bg_color="black", width=340, text_color="white")
path_label.grid(column=1, row=0, padx=10, pady=10)


file_select_btn = CTkButton(path_frame, text="Select file",height=30, width=120, command=file_select_btn_clb, font=('Helvetica', 15), text_color="white")
file_select_btn.grid(column=2, row=0, padx=10, pady=10)

mode_button = CTkButton(path_frame, text="Light", width=40, height=40, fg_color="white", hover_color="white", text_color="black", font=('Helvetica', 15), command=costumize)
mode_button.grid(row=0, column=3, padx=10, pady=10)

# ---------------------------------------------------------------------------------

# ------------ Content Sheet Frame -------------------------------------------------------

sheet_label = CTkLabel(sheet_frame, text="Sheet", font=('Helvetica', 18), text_color="white")
sheet_label.grid(column=0, row=0, padx=10, pady=10)

combobox = CTkComboBox(sheet_frame, variable=drop_down_sheet_list)
combobox.grid(column=1, row=0, padx=10, pady=10)
combobox.configure(state=DISABLED)

sheets_btn = CTkButton(sheet_frame, text="Load Sheets", command=load_sheets_clb, font=('Helvetica', 15), text_color="white")
sheets_btn.grid(row=0, column=2, padx=30, pady=10)
sheets_btn.configure(state=DISABLED)

template_btn = CTkButton(sheet_frame, text="Create Template", font=('Helvetica', 15), text_color="white", cursor = "hand2", command=template_btn_clb)
template_btn.grid(row=0, column=3)


# -----------------------------------------------------------------------------------

# ------------ Content Date Frame ----------------------------------------------

email_label = CTkLabel(date_frame, text="Email", font=('Helvetica', 18), text_color="white")
email_label.grid(column=0, row=0, padx=0, pady=30)

password_label = CTkLabel(date_frame, text="Password", font=('Helvetica', 18))
password_label.grid(column=2, row=0, padx=10, pady=10)

subject_label = CTkLabel(date_frame, text="Subject", font=('Helvetica', 18), text_color="white")
subject_label.grid(row=2, column=0, padx=10, pady=10)

input_subject = CTkEntry(date_frame)
input_subject.grid(row=2, column=1, padx=10, pady=10, sticky="w")

rememberMe_checkbutton = CTkCheckBox(date_frame, text="Remember me", variable=remember_var, onvalue=1, offvalue=0, font=('Helvetica', 14), text_color="white")
rememberMe_checkbutton.grid(row=1, column=2, columnspan=2)


# ------------------------------------------------------------------------------

# ---------- Content Checkbutton Frame ------------------------------------------

grades_checkbutton = CTkCheckBox(checkbuttons_frame, text="Grades", variable=grades_var, onvalue=1, offvalue=0, font=('Helvetica', 16), text_color="white")
grades_checkbutton.grid(row=0, column=0, padx=10, pady=20)
grades_checkbutton.select()

absente_checkbutton = CTkCheckBox(checkbuttons_frame, text="Absente", variable=absente_var, onvalue=1, offvalue=0, font=('Helvetica', 16), text_color="white")
absente_checkbutton.grid(row=0, column=1, padx=10, pady=20)

extraText_checkbutton = CTkCheckBox(checkbuttons_frame, text="Extra text", variable=extra_text_var, onvalue=1, offvalue=0, command=extraText_callback, font=('Helvetica', 16), text_color="white")
extraText_checkbutton.grid(row=0, column=2, padx=10, pady=20)

restante_checkbutton = CTkCheckBox(checkbuttons_frame, text="Restante", variable=restante_var, onvalue=1, offvalue=0, command=restante_callback, font=('Helvetica', 16), text_color="white")
restante_checkbutton.grid(row=0, column=3, padx=10, pady=20)

# -------------------------------------------------------------------------------

# ------------- Content Restante_Frame--------------------------------------------

room_label = CTkLabel(restante_frame, text="Room", font=('Helvetica', 16), text_color="white")
room_label.grid(row=0, column=0, padx=10, pady=10)

input_room = CTkEntry(restante_frame, state=DISABLED)
input_room.grid(row=0, column=1, padx=10, pady=10)

date_label = CTkLabel(restante_frame, text="Exam date", font=('Helvetica', 16), text_color="white")
date_label.grid(row=0, column=2, padx=10, pady=10)

input_date = CTkEntry(restante_frame, state=DISABLED)
input_date.grid(row=0, column=3, padx=10, pady=10)

# --------------------------------------------------------------------------------

# -------------------- Content Buttons Frame -------------------------------------

get_data_btn = CTkButton(buttons_frame, text="Get students data", command=get_data_btn_clb, font=('Helvetica', 15), text_color="white")
get_data_btn.grid(row=0, column=0, padx=10, pady=10)

parameters_selected = CTkButton(buttons_frame, text="Send", width=100, command=main, font=('Helvetica', 15), text_color="white")
parameters_selected.grid(row=0, column=1, padx=10, pady=10)
parameters_selected.configure(state=DISABLED)

error_label = CTkLabel(buttons_frame, text="Complete all the \nentries, please!", text_color="red")
# error_label.grid(row=0, column=2, padx=10)

button_info = CTkButton(buttons_frame, text="Info password", cursor = "hand2", font=('Helvetica', 15), text_color="white", command=button_info_clb)
button_info.place(x=430, y=10)

# ------------------------------------------------------------------------------

# --------------------- Content TextBox Frame ------------------------------------

info_students_data = CTkTextbox(textBox_Frame, height=70, width=350, state = "disabled");
info_students_data.insert(1.0, "")
info_students_data.grid(row=0, column=0, padx=10, pady=10, sticky="w")

text_proba = CTkTextbox(textBox_Frame, height=100, width=350)
text_proba.insert(1.0, "")
# info_students_data.grid(row=1, column=0, padx=10, pady=10, sticky="w")

frame_info = CTkFrame(textBox_Frame, height=70, border_width=2, border_color="white")


# ------- Content Frame Info ---------------------------

label_info = CTkLabel(frame_info, text="Press the button to view the tutorial \n (0:33 - 3:00 min)", font=('Helvetica', 13), text_color="white")
label_info.grid(row=0, column=0, padx=10, pady=10)

tutorial_button = CTkButton(frame_info, text="Tutorial", command=tutorial_button_info_clb)
tutorial_button.grid(row=1, column=0, padx=5, pady=5)

# ------------------------------------------------------------------------------


# --------------------- Content progressBar Frame ------------------------------------

progress_bar = ttk.Progressbar(progressBar_frame, orient=HORIZONTAL, length=400, mode="determinate")

progress_bar_label = CTkLabel(progressBar_frame, text='', font=('Helvetica', 15), text_color="white")

# ---------------------------------------------------------------------------------

# ------------ Content Contact details Frame --------------------------------------------

details_author_label = CTkLabel(contact_details_frame, text="Made by Alexandre Urluescu, contact: alexurluescu23@gmail.com", font=('Helvetica', 15), text_color="white")
details_author_label.pack()

root.mainloop()

if __name__ == "__main__":
    main()
    

